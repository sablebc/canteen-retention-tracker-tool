"""Import a Retention Tracker .xlsx export into the retention models.

Usage::

    python manage.py ingest_tracker <path-to-xlsx> [--dry-run]

The tracker is a hand-maintained spreadsheet, so the sheet carries a
status-count legend above the data and interleaves parent-company header rows
between the real site rows. Layout expectations are captured in the module
constants below and validated at run time rather than assumed.

Re-running is safe: Site, RevenueSnapshot, RepAssignment, and CallRecord are
upserted on ``site_id`` (plus the ingestion date for revenue), so a second run
over the same export updates in place instead of duplicating. StatusHistory is
the deliberate exception — it is append-only, gaining one row per site per run
so status changes stay queryable as a trend.
"""
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from retention.ingest_normalizers import (
    clean_text,
    is_ambiguous_date,
    is_usable_site_id,
    is_valid_rep_initials,
    make_synthetic_site_id,
    normalize_lob,
    normalize_site_id,
    normalize_status,
    parse_date,
    parse_decimal,
    parse_duration_minutes,
    parse_rating,
    truncate,
)
from retention.models import (
    CallRecord,
    RepAssignment,
    RevenueSnapshot,
    Site,
    StatusHistory,
)

SHEET_NAME = "Retention Tracker"
HEADER_ROW = 1
# Rows 2-10 are a status-count legend and row 11 is blank; real sites start at 12.
DATA_START_ROW = 12
# Cap on how many example values are echoed per warning category.
MAX_REPORTED_EXAMPLES = 10

# Logical field -> accepted header spellings. Comparison is done on a squashed
# form (lowercase, alphanumerics and '#' only) because the real export drifts
# in punctuation and trailing spaces — "How Long did the call take ?" has a
# space before the question mark. Q1-Q4 are matched by prefix instead, since
# their headers are long questions that get reworded between exports.
_HEADER_ALIASES = {
    "rep_initials": ("rs",),
    "site_id": ("#", "no", "site#", "siteno"),
    "name": ("sitename", "site", "name"),
    "address": ("address",),
    "method_of_ordering": ("methodofordering", "methodoforder"),
    "contact_name": ("contactname", "contact"),
    "branch": ("branch",),
    "f25_revenue": ("f25", "f25revenue"),
    "annual_revenue": ("annualrevenue", "annualrev"),
    "lob": ("lob", "lineofbusiness"),
    "phone_number": ("phonenumber", "phone", "telephone"),
    "account_status": ("accountstatus", "status"),
    "last_order_date": ("lastorderdate", "lastorder"),
    "duration_minutes": ("howlongdidthecalltake", "calllength", "duration"),
    "notes": ("notesfromcall", "notes"),
    "actions_required": ("actionsrequired", "actionrequired", "actions"),
}
_HEADER_SQUASH = re.compile(r"[^a-z0-9#]")
_QUESTION_PREFIXES = {"q1": "q1", "q2": "q2", "q3": "q3", "q4": "q4"}
_REQUIRED_FIELDS = ("site_id", "name")


@dataclass(frozen=True)
class TrackerRow:
    """One parsed site row, already normalised and ready to persist."""

    row_number: int
    site_id: str
    has_native_id: bool
    name: str
    address: str
    branch: str
    lob: str
    phone_number: str
    contact_name: str
    method_of_ordering: str
    account_status: str
    raw_account_status: str
    last_order_date: date | None
    f25_revenue: Decimal | None
    annual_revenue: Decimal | None
    rep_initials: str
    rating: int | None
    q1: str
    q2: str
    q4: str
    notes: str
    actions_required: str
    duration_minutes: int | None

    @property
    def has_call_content(self) -> bool:
        """True when the row records an actual conversation.

        A row carrying only "Actions required" is an attempt log, not a call,
        and must not produce an empty CallRecord.
        """
        return bool(
            self.rating is not None or self.q1 or self.q2 or self.q4 or self.notes
        )


@dataclass
class IngestStats:
    """Mutable counters describing one parse/load pass."""

    rows_scanned: int = 0
    preamble_skipped: int = 0
    blank_skipped: int = 0
    group_headers_skipped: int = 0
    group_header_names: list[str] = field(default_factory=list)
    synthetic_id_rows: int = 0
    native_id_rows: int = 0
    named_no_name_with_id: int = 0
    invalid_rep_values: list[str] = field(default_factory=list)
    rejected_native_ids: list[str] = field(default_factory=list)
    synthetic_collisions: list[str] = field(default_factory=list)
    rep_breakdown: Counter = field(default_factory=Counter)
    sites_created: int = 0
    sites_updated: int = 0
    revenue_created: int = 0
    revenue_updated: int = 0
    revenue_skipped_empty: int = 0
    reps_created: int = 0
    reps_updated: int = 0
    reps_skipped_blank: int = 0
    calls_created: int = 0
    calls_updated: int = 0
    calls_skipped_empty: int = 0
    status_history_appended: int = 0
    duplicate_site_ids: list[str] = field(default_factory=list)
    unparseable_dates: list[str] = field(default_factory=list)
    ambiguous_dates: list[str] = field(default_factory=list)
    bad_ratings: list[str] = field(default_factory=list)
    unparseable_durations: list[str] = field(default_factory=list)
    status_breakdown: dict[str, int] = field(default_factory=dict)

    def note(self, bucket: list[str], value: str) -> None:
        """Record an example value, keeping only the first few per category."""
        if len(bucket) < MAX_REPORTED_EXAMPLES:
            bucket.append(value)


def _normalize_header(value) -> str:
    """Squash a header cell to a punctuation-insensitive key, preserving '#'."""
    return _HEADER_SQUASH.sub("", clean_text(value).lower())


def resolve_columns(header_cells: tuple) -> dict[str, int]:
    """Map logical field names onto zero-based column indexes.

    Matching is by header text rather than fixed column letters so an inserted
    column does not silently shift every field.

    Raises:
        CommandError: If a required header ('#', 'Site Name') is absent.
    """
    columns: dict[str, int] = {}
    for index, cell in enumerate(header_cells):
        header = _normalize_header(cell)
        if not header:
            continue

        for key, prefix in _QUESTION_PREFIXES.items():
            if header.startswith(prefix) and key not in columns:
                columns[key] = index
                break
        else:
            for field_name, aliases in _HEADER_ALIASES.items():
                if header in aliases and field_name not in columns:
                    columns[field_name] = index
                    break

    missing = [name for name in _REQUIRED_FIELDS if name not in columns]
    if missing:
        raise CommandError(
            f"Sheet '{SHEET_NAME}' is missing required column(s): "
            f"{', '.join(missing)}. Found headers: "
            f"{[_normalize_header(c) for c in header_cells if _normalize_header(c)]}"
        )
    return columns


def _cell(row: tuple, columns: dict[str, int], field_name: str):
    """Return a raw cell value by logical field name, or None if absent."""
    index = columns.get(field_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def build_row(row: tuple, columns: dict[str, int], row_number: int,
              stats: IngestStats, native_id: str) -> TrackerRow:
    """Normalise one spreadsheet row into a TrackerRow, recording any oddities.

    ``native_id`` is the already-validated '#' value (empty when the row needs
    a synthetic ID), passed in so classification and construction cannot
    disagree about which key a row owns.
    """
    raw_date = _cell(row, columns, "last_order_date")
    last_order_date, date_failed = parse_date(raw_date)
    if date_failed:
        stats.note(stats.unparseable_dates, f"row {row_number}: {raw_date!r}")
    elif last_order_date and is_ambiguous_date(raw_date):
        stats.note(stats.ambiguous_dates, f"row {row_number}: {raw_date!r}")

    raw_rating = _cell(row, columns, "q3")
    rating, rating_bad = parse_rating(raw_rating)
    if rating_bad:
        stats.note(stats.bad_ratings, f"row {row_number}: {raw_rating!r}")

    raw_duration = _cell(row, columns, "duration_minutes")
    duration, duration_bad = parse_duration_minutes(raw_duration)
    if duration_bad:
        stats.note(stats.unparseable_durations, f"row {row_number}: {raw_duration!r}")

    raw_status = clean_text(_cell(row, columns, "account_status"))
    status = normalize_status(raw_status)
    stats.status_breakdown[status] = stats.status_breakdown.get(status, 0) + 1

    name = clean_text(_cell(row, columns, "name"))
    address = clean_text(_cell(row, columns, "address"))
    rep_initials = clean_text(_cell(row, columns, "rep_initials"))
    if not is_valid_rep_initials(rep_initials):
        rep_initials = ""

    return TrackerRow(
        row_number=row_number,
        site_id=native_id or make_synthetic_site_id(name, address),
        has_native_id=bool(native_id),
        name=name,
        address=address,
        branch=clean_text(_cell(row, columns, "branch")),
        lob=normalize_lob(_cell(row, columns, "lob")),
        phone_number=clean_text(_cell(row, columns, "phone_number")),
        contact_name=clean_text(_cell(row, columns, "contact_name")),
        method_of_ordering=clean_text(_cell(row, columns, "method_of_ordering")),
        account_status=status,
        raw_account_status=raw_status,
        last_order_date=last_order_date,
        f25_revenue=parse_decimal(_cell(row, columns, "f25_revenue")),
        annual_revenue=parse_decimal(_cell(row, columns, "annual_revenue")),
        rep_initials=rep_initials,
        rating=rating,
        q1=clean_text(_cell(row, columns, "q1")),
        q2=clean_text(_cell(row, columns, "q2")),
        q4=clean_text(_cell(row, columns, "q4")),
        notes=clean_text(_cell(row, columns, "notes")),
        actions_required=clean_text(_cell(row, columns, "actions_required")),
        duration_minutes=duration,
    )


def parse_workbook(path: Path, stats: IngestStats) -> list[TrackerRow]:
    """Read the tracker sheet and return its real site rows.

    Rows are classified as: preamble (the legend block above the data), a site
    row (column '#' populated), a parent-company group header (a name but no
    '#'), or blank.

    Raises:
        CommandError: If the file or sheet cannot be read.
    """
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a variety of types here.
        raise CommandError(f"Could not open workbook '{path}': {exc}") from exc

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise CommandError(
                f"Sheet '{SHEET_NAME}' not found. Available sheets: "
                f"{workbook.sheetnames}"
            )
        sheet = workbook[SHEET_NAME]

        rows: list[TrackerRow] = []
        seen_ids: set[str] = set()
        columns: dict[str, int] = {}

        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number == HEADER_ROW:
                columns = resolve_columns(row)
                continue

            stats.rows_scanned += 1
            if row_number < DATA_START_ROW:
                stats.preamble_skipped += 1
                continue

            native_id = normalize_site_id(_cell(row, columns, "site_id"))
            # Any '#' text marks the row as data, but only a value containing a
            # digit is usable as a key — the sheet has '****' on two unrelated
            # rows, and honouring it would merge them into one site.
            has_id_text = bool(native_id)
            if native_id and not is_usable_site_id(native_id):
                stats.note(
                    stats.rejected_native_ids, f"row {row_number}: {native_id!r}"
                )
                native_id = ""

            name = clean_text(_cell(row, columns, "name"))
            rep = clean_text(_cell(row, columns, "rep_initials"))
            has_rep = is_valid_rep_initials(rep)
            if rep and not has_rep:
                stats.note(stats.invalid_rep_values, f"row {row_number}: {rep!r}")

            # A row is a real site when it is named AND carries either the
            # sheet's own '#' or a rep assignment. Requiring '#' alone dropped
            # hundreds of RS-tagged sites into the skipped buckets.
            if not name:
                if native_id:
                    stats.named_no_name_with_id += 1
                else:
                    stats.blank_skipped += 1
                continue

            if not native_id and not has_rep:
                # Named, unassigned, no '#': a parent-company header for the
                # rows beneath it. Counted so it can later backfill children.
                stats.group_headers_skipped += 1
                stats.note(stats.group_header_names, f"row {row_number}: {name}")
                continue

            entry = build_row(row, columns, row_number, stats, native_id)
            if entry.has_native_id:
                stats.native_id_rows += 1
            else:
                stats.synthetic_id_rows += 1

            if entry.site_id in seen_ids:
                bucket = (
                    stats.synthetic_collisions
                    if not entry.has_native_id
                    else stats.duplicate_site_ids
                )
                stats.note(bucket, f"row {row_number}: {name} -> {entry.site_id}")
            seen_ids.add(entry.site_id)
            rows.append(entry)

        return rows
    finally:
        workbook.close()


def _upsert_site(entry: TrackerRow, stats: IngestStats) -> Site:
    """Create or update the Site keyed on its tracker '#'."""
    site, created = Site.objects.update_or_create(
        site_id=entry.site_id,
        defaults={
            "has_native_id": entry.has_native_id,
            "name": entry.name,
            "address": entry.address,
            "branch": entry.branch,
            "lob": entry.lob,
            "phone_number": entry.phone_number,
            "contact_name": entry.contact_name,
            "method_of_ordering": entry.method_of_ordering,
            "account_status": entry.account_status,
            "raw_account_status": truncate(entry.raw_account_status, 255),
            "last_order_date": entry.last_order_date,
        },
    )
    if created:
        stats.sites_created += 1
    else:
        stats.sites_updated += 1
    return site


def _upsert_revenue(site: Site, entry: TrackerRow, snapshot_date: date,
                    stats: IngestStats) -> None:
    """Record today's revenue reading, one snapshot per site per ingestion day."""
    if entry.f25_revenue is None and entry.annual_revenue is None:
        stats.revenue_skipped_empty += 1
        return

    _, created = RevenueSnapshot.objects.update_or_create(
        site=site,
        snapshot_date=snapshot_date,
        defaults={
            "f25_revenue": entry.f25_revenue,
            "annual_revenue": entry.annual_revenue,
        },
    )
    if created:
        stats.revenue_created += 1
    else:
        stats.revenue_updated += 1


def _upsert_rep(site: Site, entry: TrackerRow, stats: IngestStats) -> None:
    """Assign the rep only when RS holds a real value; blank means unassigned."""
    if not entry.rep_initials:
        stats.reps_skipped_blank += 1
        return

    _, created = RepAssignment.objects.update_or_create(
        site=site,
        defaults={"rep_initials": truncate(entry.rep_initials, 16)},
    )
    stats.rep_breakdown[entry.rep_initials] += 1
    if created:
        stats.reps_created += 1
    else:
        stats.reps_updated += 1


def _upsert_call(site: Site, entry: TrackerRow, stats: IngestStats) -> None:
    """Persist the row's call block, skipping rows with no real conversation.

    The tracker holds at most one call block per site and carries no call date,
    so (site, call_date=None) is the natural upsert key for re-runs.
    """
    if not entry.has_call_content:
        stats.calls_skipped_empty += 1
        return

    defaults = {
        "source": CallRecord.Source.IMPORT,
        "rep_initials": truncate(entry.rep_initials, 16),
        "duration_minutes": entry.duration_minutes,
        "q1_last_order_feedback": entry.q1,
        "q2_working_well": entry.q2,
        "rating": entry.rating,
        "q4_could_improve": entry.q4,
        "notes": entry.notes,
        "actions_required": entry.actions_required,
    }
    existing = CallRecord.objects.filter(site=site, call_date=None).order_by("pk")
    first = existing.first()
    if first is None:
        CallRecord.objects.create(site=site, call_date=None, **defaults)
        stats.calls_created += 1
        return

    for attribute, value in defaults.items():
        setattr(first, attribute, value)
    first.save(update_fields=list(defaults))
    stats.calls_updated += 1


class Command(BaseCommand):
    """Management command entry point."""

    help = "Import a Retention Tracker .xlsx export into the retention models."

    def add_arguments(self, parser) -> None:
        parser.add_argument("xlsx_path", type=str, help="Path to the .xlsx export.")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report counts without persisting anything.",
        )

    def handle(self, *args, **options) -> None:
        path = Path(options["xlsx_path"]).expanduser()
        dry_run = options["dry_run"]

        if not path.is_file():
            raise CommandError(f"File not found: {path}")

        stats = IngestStats()
        rows = parse_workbook(path, stats)
        observed_date = timezone.localdate()

        # The dry run executes the real write path inside a transaction that is
        # always rolled back, so the reported counts are exact rather than
        # predicted, and integrity errors surface before a live run.
        with transaction.atomic():
            self._load(rows, stats, observed_date)
            if dry_run:
                transaction.set_rollback(True)

        self._report(stats, path, observed_date, dry_run)

    def _load(self, rows: list[TrackerRow], stats: IngestStats,
              observed_date: date) -> None:
        """Persist every parsed row and append this run's status history."""
        history: list[StatusHistory] = []
        raw_status_limit = StatusHistory._meta.get_field("raw_status").max_length
        status_limit = StatusHistory._meta.get_field("status").max_length

        for entry in rows:
            site = _upsert_site(entry, stats)
            _upsert_revenue(site, entry, observed_date, stats)
            _upsert_rep(site, entry, stats)
            _upsert_call(site, entry, stats)

            # Append-only: never update_or_create here.
            history.append(
                StatusHistory(
                    site=site,
                    status=truncate(entry.account_status, status_limit),
                    raw_status=truncate(entry.raw_account_status, raw_status_limit),
                    observed_date=observed_date,
                )
            )

        StatusHistory.objects.bulk_create(history)
        stats.status_history_appended = len(history)

    def _report(self, stats: IngestStats, path: Path, observed_date: date,
                dry_run: bool) -> None:
        """Print a human-readable summary of the run."""
        mode = "DRY RUN (nothing written)" if dry_run else "COMMITTED"
        verb = "would be" if dry_run else ""

        self.stdout.write(self.style.MIGRATE_HEADING(f"\nIngest tracker — {mode}"))
        self.stdout.write(f"  Source: {path}")
        self.stdout.write(f"  Observed date: {observed_date}")

        self.stdout.write(self.style.MIGRATE_HEADING("\nRows"))
        self.stdout.write(f"  Scanned (excl. header):   {stats.rows_scanned}")
        self.stdout.write(f"  Legend/preamble skipped:  {stats.preamble_skipped}")
        self.stdout.write(f"  Group headers skipped:    {stats.group_headers_skipped}")
        self.stdout.write(f"  Blank rows skipped:       {stats.blank_skipped}")
        if stats.named_no_name_with_id:
            self.stdout.write(
                f"  '#' present but unnamed:  {stats.named_no_name_with_id}"
            )
        self.stdout.write(
            f"  Site rows found:          {stats.sites_created + stats.sites_updated}"
        )
        self.stdout.write(f"    - with sheet '#' ID:    {stats.native_id_rows}")
        self.stdout.write(
            self.style.SUCCESS(
                f"    - Named/assigned rows recovered via synthetic ID: "
                f"{stats.synthetic_id_rows}"
            )
        )

        self.stdout.write(self.style.MIGRATE_HEADING(f"\nRecords {verb}".rstrip()))
        for label, created, updated in (
            ("Site", stats.sites_created, stats.sites_updated),
            ("RevenueSnapshot", stats.revenue_created, stats.revenue_updated),
            ("RepAssignment", stats.reps_created, stats.reps_updated),
            ("CallRecord", stats.calls_created, stats.calls_updated),
        ):
            self.stdout.write(
                f"  {label:<16} created={created:<5} updated={updated}"
            )
        self.stdout.write(
            f"  {'StatusHistory':<16} appended={stats.status_history_appended} "
            f"(append-only)"
        )

        self.stdout.write(self.style.MIGRATE_HEADING("\nSkipped sub-records"))
        self.stdout.write(f"  Rep blank (unassigned):   {stats.reps_skipped_blank}")
        self.stdout.write(f"  Call rows with no content:{stats.calls_skipped_empty}")
        self.stdout.write(f"  Rows with no revenue:     {stats.revenue_skipped_empty}")

        if stats.rep_breakdown:
            self.stdout.write(self.style.MIGRATE_HEADING("\nRepAssignment per rep"))
            for rep, count in sorted(stats.rep_breakdown.items()):
                self.stdout.write(f"  {rep:<20} {count}")
            self.stdout.write(f"  {'TOTAL':<20} {sum(stats.rep_breakdown.values())}")

        if stats.status_breakdown:
            self.stdout.write(self.style.MIGRATE_HEADING("\nNormalized statuses"))
            for status, count in sorted(stats.status_breakdown.items()):
                self.stdout.write(f"  {status:<20} {count}")

        self._report_warnings(stats)

        if dry_run:
            self.stdout.write(
                self.style.WARNING("\nDry run — transaction rolled back.\n")
            )
        else:
            self.stdout.write(self.style.SUCCESS("\nIngest complete.\n"))

    def _report_warnings(self, stats: IngestStats) -> None:
        """Surface data-quality issues worth a human look."""
        sections = (
            ("Unparseable dates (field left null)", stats.unparseable_dates),
            ("Ambiguous dates (read as day-first)", stats.ambiguous_dates),
            ("Ratings outside 1-5 (left null)", stats.bad_ratings),
            ("Unparseable call durations", stats.unparseable_durations),
            ("Duplicate '#' site IDs (last row wins)", stats.duplicate_site_ids),
            (
                "Synthetic ID collisions (identical name+address, merged)",
                stats.synthetic_collisions,
            ),
            ("RS values not recognised as rep initials", stats.invalid_rep_values),
            (
                "'#' values ignored as placeholders (synthetic ID used instead)",
                stats.rejected_native_ids,
            ),
            ("Group headers skipped", stats.group_header_names),
        )
        reported = [(title, values) for title, values in sections if values]
        if not reported:
            return

        self.stdout.write(self.style.MIGRATE_HEADING("\nWarnings"))
        for title, values in reported:
            self.stdout.write(self.style.WARNING(f"  {title}: {len(values)}"))
            for value in values[:MAX_REPORTED_EXAMPLES]:
                self.stdout.write(f"      {value}")
            if len(values) == MAX_REPORTED_EXAMPLES:
                self.stdout.write("      … (further occurrences not listed)")
